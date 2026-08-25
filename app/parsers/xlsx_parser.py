"""XLSX parser with sheet, table-region, and cell-range traceability."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from app.models.document_models import (
    DocumentBlock,
    DocumentSourceFile,
    NormalizedDocument,
)
from app.parsers.document_parser import (
    DocumentParseError,
    build_normalized_document,
)


_MAX_SCANNED_CELLS = 1_000_000


@dataclass(frozen=True)
class _Region:
    row_start: int
    row_end: int
    column_start: int
    column_end: int
    rows: tuple[tuple[str, ...], ...]


@dataclass
class _PendingRegion:
    row_start: int
    row_end: int
    columns: set[int]
    rows: list[dict[int, str]]


class XlsxDocumentParser:
    def __init__(self, *, extract_charts: bool = False) -> None:
        self._extract_charts = extract_charts

    def parse(self, source: DocumentSourceFile) -> NormalizedDocument:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise DocumentParseError(
                "openpyxl is required to parse XLSX documents"
            ) from exc

        formula_book = None
        value_book = None
        chart_book = None
        try:
            formula_book = load_workbook(
                source.path,
                read_only=True,
                data_only=False,
            )
            value_book = load_workbook(
                source.path,
                read_only=True,
                data_only=True,
            )
            if self._extract_charts:
                chart_book = load_workbook(
                    source.path,
                    read_only=False,
                    data_only=False,
                )
            blocks = self._blocks(
                formula_book,
                value_book,
                get_column_letter,
                chart_book=chart_book,
            )
            properties = formula_book.properties
            title = _string(properties.title)
            created = _date_string(properties.created)
        except DocumentParseError:
            raise
        except Exception as exc:
            raise DocumentParseError(
                f"Unable to parse XLSX: {source.path}"
            ) from exc
        finally:
            if formula_book is not None:
                formula_book.close()
            if value_book is not None:
                value_book.close()
            if chart_book is not None:
                chart_book.close()

        return build_normalized_document(
            source,
            blocks,
            detected_title=title,
            detected_created_date=created,
        )

    def _blocks(
        self,
        formula_book: Any,
        value_book: Any,
        get_column_letter: Any,
        *,
        chart_book: Any | None = None,
    ) -> list[DocumentBlock]:
        blocks: list[DocumentBlock] = []
        for formula_sheet in formula_book.worksheets:
            value_sheet = value_book[formula_sheet.title]
            regions = _sheet_regions(formula_sheet, value_sheet)
            charts = (
                tuple(chart_book[formula_sheet.title]._charts)
                if chart_book is not None
                else ()
            )
            if not regions and not charts:
                continue
            blocks.append(
                DocumentBlock(
                    "heading",
                    formula_sheet.title,
                    level=1,
                    sheet=formula_sheet.title,
                )
            )
            for region in regions:
                column_start = region.column_start
                column_end = region.column_end
                cell_range = (
                    f"{get_column_letter(column_start)}{region.row_start}:"
                    f"{get_column_letter(column_end)}{region.row_end}"
                )
                blocks.append(
                    DocumentBlock(
                        "table",
                        _format_rows(region.rows),
                        sheet=formula_sheet.title,
                        cell_range=cell_range,
                        rows=region.rows,
                    )
                )
            for chart_index, chart in enumerate(charts, 1):
                blocks.extend(
                    _chart_blocks(chart, chart_index, formula_sheet.title)
                )
        return blocks


def _chart_blocks(chart: Any, chart_index: int, sheet: str) -> list[DocumentBlock]:
    title = _rich_text(getattr(chart, "title", None)) or f"Chart {chart_index}"
    lines = [f"Chart type: {type(chart).__name__}", f"Title: {title}"]
    for label, attribute in (
        ("X axis", "x_axis"),
        ("Y axis", "y_axis"),
        ("Z axis", "z_axis"),
    ):
        axis = getattr(chart, attribute, None)
        axis_title = _rich_text(getattr(axis, "title", None))
        if axis_title:
            lines.append(f"{label}: {axis_title}")

    references: list[str] = []
    for series_index, series in enumerate(getattr(chart, "ser", ()), 1):
        series_title = _series_title(series)
        if series_title:
            lines.append(f"Series {series_index} title: {series_title}")
        for label, reference in _series_references(series):
            lines.append(f"Series {series_index} {label}: {reference}")
            references.append(reference)

    return [
        DocumentBlock("heading", f"Chart: {title}", level=2, sheet=sheet),
        DocumentBlock(
            "paragraph",
            "\n".join(lines),
            sheet=sheet,
            cell_range="; ".join(dict.fromkeys(references)),
        ),
    ]


def _series_title(series: Any) -> str:
    title = getattr(series, "tx", None)
    value = getattr(title, "v", None)
    if isinstance(value, str) and value.strip():
        return value.strip()
    reference = _reference_formula(title)
    return reference or ""


def _series_references(series: Any) -> list[tuple[str, str]]:
    references: list[tuple[str, str]] = []
    for label, attribute in (
        ("categories", "cat"),
        ("values", "val"),
        ("X values", "xVal"),
        ("Y values", "yVal"),
        ("Z values", "zVal"),
    ):
        formula = _reference_formula(getattr(series, attribute, None))
        if formula:
            references.append((label, formula))
    return references


def _reference_formula(value: Any) -> str:
    if value is None:
        return ""
    for attribute in ("numRef", "strRef"):
        reference = getattr(value, attribute, None)
        formula = getattr(reference, "f", None)
        if isinstance(formula, str) and formula.strip():
            return formula.strip()
    return ""


def _rich_text(value: Any, *, _seen: set[int] | None = None, _depth: int = 0) -> str:
    if value is None or _depth > 8:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (tuple, list)):
        return " ".join(
            item for child in value if (item := _rich_text(child, _seen=_seen, _depth=_depth + 1))
        ).strip()
    seen = _seen if _seen is not None else set()
    identity = id(value)
    if identity in seen:
        return ""
    seen.add(identity)
    fragments: list[str] = []
    for attribute in ("tx", "rich", "p", "r", "t", "v"):
        if hasattr(value, attribute):
            text = _rich_text(
                getattr(value, attribute),
                _seen=seen,
                _depth=_depth + 1,
            )
            if text:
                fragments.append(text)
    return " ".join(fragments).strip()


def _sheet_regions(
    formula_sheet: Any,
    value_sheet: Any,
) -> list[_Region]:
    cell_count = formula_sheet.max_row * formula_sheet.max_column
    if cell_count > _MAX_SCANNED_CELLS:
        raise DocumentParseError(
            f"XLSX sheet is too large to scan safely: {formula_sheet.title}"
        )

    populated: list[tuple[int, dict[int, str]]] = []
    formula_rows = formula_sheet.iter_rows()
    value_rows = value_sheet.iter_rows()
    for row_index, (formula_row, value_row) in enumerate(
        zip(formula_rows, value_rows),
        1,
    ):
        values = {
            column_index: text
            for column_index, (formula_cell, value_cell) in enumerate(
                zip(formula_row, value_row),
                1,
            )
            if (text := _cell_text(formula_cell, value_cell))
        }
        if values:
            populated.append((row_index, values))

    grouped: list[_PendingRegion] = []
    for row_index, values in populated:
        columns = set(values)
        if grouped and row_index == grouped[-1].row_end + 1:
            grouped[-1].row_end = row_index
            grouped[-1].columns.update(columns)
            grouped[-1].rows.append(values)
        else:
            grouped.append(
                _PendingRegion(row_index, row_index, columns, [values])
            )

    regions: list[_Region] = []
    for pending in grouped:
        column_start = min(pending.columns)
        column_end = max(pending.columns)
        rows = tuple(
            tuple(
                values.get(column_index, "")
                for column_index in range(column_start, column_end + 1)
            )
            for values in pending.rows
        )
        regions.append(
            _Region(
                pending.row_start,
                pending.row_end,
                column_start,
                column_end,
                rows,
            )
        )
    return regions


def _cell_text(formula_cell: Any, value_cell: Any) -> str:
    value = formula_cell.value
    if formula_cell.data_type == "f" or (
        isinstance(value, str) and value.startswith("=")
    ):
        formula = value if str(value).startswith("=") else f"={value}"
        cached = _scalar_text(value_cell.value)
        if cached:
            return f"Formula: {formula}; Cached value: {cached}"
        return f"Formula: {formula}"
    return _scalar_text(value)


def _scalar_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    return str(value).strip()


def _format_rows(rows: tuple[tuple[str, ...], ...]) -> str:
    return "\n".join(" | ".join(row) for row in rows)


def _string(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def _date_string(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return ""
